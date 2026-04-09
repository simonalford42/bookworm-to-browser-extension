"""
Plots for the Yanus & Ardoin (2026) replication + age analysis
"""
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from collections import defaultdict

# ── Load data ─────────────────────────────────────────────────────────────────
print("Loading data...")
wb = openpyxl.load_workbook('Journal_Data_Final.xlsx', read_only=True)
ws = wb.active
next(ws.iter_rows(min_row=1, max_row=1))  # skip header

refs = []
for row in ws.iter_rows(min_row=2, values_only=True):
    ref_type = row[4]
    year     = row[6]
    pub_year = row[8]
    subfield = row[10]
    if year and ref_type:
        refs.append((ref_type, year, pub_year, subfield))

def classify(ref_type):
    if ref_type in ('Book', 'Book Section'):
        return 'book'
    elif ref_type == 'Journal Article':
        return 'article'
    return 'other'

YEARS = [1990, 1995, 2000, 2005, 2010, 2015, 2020, 2024]

# ── Pre-compute by-year counts ─────────────────────────────────────────────
by_year = defaultdict(lambda: defaultdict(int))
for ref_type, year, pub_year, subfield in refs:
    cat = classify(ref_type)
    by_year[year][cat] += 1

art_pct  = [100 * by_year[y]['article'] / sum(by_year[y].values()) for y in YEARS]
book_pct = [100 * by_year[y]['book']    / sum(by_year[y].values()) for y in YEARS]
other_pct= [100 * by_year[y]['other']   / sum(by_year[y].values()) for y in YEARS]

# ── Pre-compute by-subfield-year ───────────────────────────────────────────
SUBFIELDS = ['AmerPol', 'CompPol', 'IR', 'PolTheory', 'PubAdm']
SF_LABELS  = ['American Politics', 'Comparative Politics', 'Intl Relations',
               'Political Theory', 'Public Admin']

by_sf_year = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
for ref_type, year, pub_year, subfield in refs:
    if subfield and subfield != '':
        cat = classify(ref_type)
        by_sf_year[subfield][year][cat] += 1

# ── Pre-compute ages ───────────────────────────────────────────────────────
age_by_cat = defaultdict(list)
age_by_year_cat = defaultdict(lambda: defaultdict(list))
for ref_type, year, pub_year, subfield in refs:
    cat = classify(ref_type)
    if cat in ('book', 'article') and pub_year and year:
        try:
            age = int(year) - int(pub_year)
            if 0 <= age <= 100:
                age_by_cat[cat].append(age)
                age_by_year_cat[year][cat].append(age)
        except (ValueError, TypeError):
            pass

mean_age_book    = [sum(age_by_year_cat[y]['book'])    / len(age_by_year_cat[y]['book'])    for y in YEARS]
mean_age_article = [sum(age_by_year_cat[y]['article']) / len(age_by_year_cat[y]['article']) for y in YEARS]

# ── Plotting ───────────────────────────────────────────────────────────────
fig, axes = plt.subplots(2, 2, figsize=(13, 10))
fig.suptitle('Political Science Citation Patterns, 1990–2024\n'
             'Data: Yanus & Ardoin (2026)', fontsize=13, y=1.01)

COLORS = {'article': '#2166ac', 'book': '#d6604d', 'other': '#999999'}
SF_COLORS = ['#1b7837', '#762a83', '#2166ac', '#d6604d', '#e08214']

# ── Plot 1: % books vs articles over time (replicates Figure 1) ────────────
ax = axes[0, 0]
ax.plot(YEARS, art_pct,  marker='o', color=COLORS['article'], linewidth=2, label='Journal Articles')
ax.plot(YEARS, book_pct, marker='s', color=COLORS['book'],    linewidth=2, label='Books')
ax.plot(YEARS, other_pct,marker='^', color=COLORS['other'],   linewidth=2, label='Other', linestyle='--')
for x, y in zip(YEARS, art_pct):
    ax.annotate(f'{y:.1f}%', (x, y), textcoords='offset points', xytext=(0, 7), ha='center', fontsize=7.5)
for x, y in zip(YEARS, book_pct):
    ax.annotate(f'{y:.1f}%', (x, y), textcoords='offset points', xytext=(0, -13), ha='center', fontsize=7.5)
ax.set_title('Figure 1 (replicated): Share of References by Type')
ax.set_ylabel('% of all references')
ax.set_ylim(0, 80)
ax.set_xticks(YEARS)
ax.legend()
ax.grid(axis='y', alpha=0.3)

# ── Plot 2: % journal articles by subfield (replicates Figure 2) ──────────
ax = axes[0, 1]
for sf, label, color in zip(SUBFIELDS, SF_LABELS, SF_COLORS):
    pcts = []
    for y in YEARS:
        d = by_sf_year[sf][y]
        n = sum(d.values())
        pcts.append(100 * d['article'] / n if n else float('nan'))
    ax.plot(YEARS, pcts, marker='o', linewidth=2, label=label, color=color)
ax.set_title('Figure 2 (replicated): % Journal Article References by Subfield')
ax.set_ylabel('% journal article references')
ax.set_ylim(0, 100)
ax.set_xticks(YEARS)
ax.legend(fontsize=8)
ax.grid(axis='y', alpha=0.3)

# ── Plot 3: age distribution histogram (new analysis) ─────────────────────
ax = axes[1, 0]
bins = range(0, 55, 2)
book_ages    = age_by_cat['book']
article_ages = age_by_cat['article']
ax.hist(book_ages,    bins=bins, density=True, alpha=0.6, color=COLORS['book'],
        label=f'Books (n={len(book_ages):,})')
ax.hist(article_ages, bins=bins, density=True, alpha=0.6, color=COLORS['article'],
        label=f'Articles (n={len(article_ages):,})')
book_mean    = sum(book_ages)    / len(book_ages)
article_mean = sum(article_ages) / len(article_ages)
ax.axvline(book_mean,    color=COLORS['book'],    linestyle='--', linewidth=1.5,
           label=f'Book mean: {book_mean:.1f} yrs')
ax.axvline(article_mean, color=COLORS['article'], linestyle='--', linewidth=1.5,
           label=f'Article mean: {article_mean:.1f} yrs')
ax.set_title('Age of Cited Works at Time of Citation')
ax.set_xlabel('Years since publication')
ax.set_ylabel('Density')
ax.legend(fontsize=8.5)
ax.grid(axis='y', alpha=0.3)
ax.set_xlim(0, 54)

# ── Plot 4: mean citation age over time, books vs articles ────────────────
ax = axes[1, 1]
ax.plot(YEARS, mean_age_book,    marker='s', color=COLORS['book'],    linewidth=2, label='Books')
ax.plot(YEARS, mean_age_article, marker='o', color=COLORS['article'], linewidth=2, label='Articles')
for x, y in zip(YEARS, mean_age_book):
    ax.annotate(f'{y:.1f}', (x, y), textcoords='offset points', xytext=(0, 7), ha='center', fontsize=8)
for x, y in zip(YEARS, mean_age_article):
    ax.annotate(f'{y:.1f}', (x, y), textcoords='offset points', xytext=(0, -13), ha='center', fontsize=8)
ax.set_title('Mean Age of Cited Work at Time of Citation')
ax.set_ylabel('Mean age (years)')
ax.set_xticks(YEARS)
ax.legend()
ax.grid(axis='y', alpha=0.3)
ax.set_ylim(0, 25)

plt.tight_layout()
plt.savefig('citation_analysis.png', dpi=150, bbox_inches='tight')
print("Saved citation_analysis.png")

# ── Mobile version: just the two new plots, stacked vertically ───────────
fig2, (ax3, ax4) = plt.subplots(2, 1, figsize=(6, 10))

# Plot 3 again
ax3.hist(book_ages,    bins=bins, density=True, alpha=0.6, color=COLORS['book'],
         label=f'Books (n={len(book_ages):,})')
ax3.hist(article_ages, bins=bins, density=True, alpha=0.6, color=COLORS['article'],
         label=f'Articles (n={len(article_ages):,})')
ax3.axvline(book_mean,    color=COLORS['book'],    linestyle='--', linewidth=1.5,
            label=f'Book mean: {book_mean:.1f} yrs')
ax3.axvline(article_mean, color=COLORS['article'], linestyle='--', linewidth=1.5,
            label=f'Article mean: {article_mean:.1f} yrs')
ax3.set_title('Age of Cited Works at Time of Citation')
ax3.set_xlabel('Years since publication')
ax3.set_ylabel('Density')
ax3.legend(fontsize=9)
ax3.grid(axis='y', alpha=0.3)
ax3.set_xlim(0, 54)

# Plot 4 again
ax4.plot(YEARS, mean_age_book,    marker='s', color=COLORS['book'],    linewidth=2, label='Books')
ax4.plot(YEARS, mean_age_article, marker='o', color=COLORS['article'], linewidth=2, label='Articles')
for x, y in zip(YEARS, mean_age_book):
    ax4.annotate(f'{y:.1f}', (x, y), textcoords='offset points', xytext=(0, 7), ha='center', fontsize=8)
for x, y in zip(YEARS, mean_age_article):
    ax4.annotate(f'{y:.1f}', (x, y), textcoords='offset points', xytext=(0, -13), ha='center', fontsize=8)
ax4.set_title('Mean Age of Cited Work at Time of Citation')
ax4.set_ylabel('Mean age (years)')
ax4.set_xticks(YEARS)
ax4.tick_params(axis='x', labelrotation=45)
ax4.legend()
ax4.grid(axis='y', alpha=0.3)
ax4.set_ylim(0, 25)

fig2.suptitle('How Old Are the Works Cited in Political Science?\n'
              'Data: Yanus & Ardoin (2026), 50,453 citations, 1990–2024',
              fontsize=11, y=1.01)
fig2.tight_layout()
fig2.savefig('citation_age_analysis.png', dpi=150, bbox_inches='tight')
print("Saved citation_age_analysis.png")
