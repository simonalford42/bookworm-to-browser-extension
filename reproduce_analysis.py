"""
Reproduce the analysis from:
"From Bookworm to Browser: The Decline of Books in Political Science Scholarship"
Yanus & Ardoin (2026), plus age-of-citation analysis (original question)
"""
import csv
import openpyxl
from collections import defaultdict

# ── Load full reference-level data ───────────────────────────────────────────
print("Loading Journal_Data_Final.xlsx ...")
wb = openpyxl.load_workbook('Journal_Data_Final.xlsx', read_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
# ID, Count, ItemType, Book, Reference, Journal, Year, Article, PublicationYear,
# Count2, subfield, comments, PublicationTitle, Title

refs = []
for row in ws.iter_rows(min_row=2, values_only=True):
    ref_type = row[4]   # 'Book', 'Book Section', 'Journal Article', etc.
    year = row[6]       # year of citing article
    pub_year = row[8]   # year cited work was published
    subfield = row[10]
    journal = row[5]
    if year and ref_type:
        refs.append({
            'ref_type': ref_type,
            'year': year,
            'pub_year': pub_year,
            'subfield': subfield,
            'journal': journal,
        })

print(f"Total references: {len(refs)}")

# Classify as book (Book + Book Section) or article (Journal Article)
def classify(ref_type):
    if ref_type in ('Book', 'Book Section'):
        return 'book'
    elif ref_type == 'Journal Article':
        return 'article'
    else:
        return 'other'

for r in refs:
    r['cat'] = classify(r['ref_type'])


# ── TABLE 1 / FIGURE 1: % books vs articles by year ────────────────────────
print("\n=== TABLE 1: References by Type, 1990–2024 ===")
print(f"{'Year':>6} {'Articles%':>10} {'Books%':>8} {'Other%':>7} {'N':>7}")

by_year = defaultdict(lambda: defaultdict(int))
for r in refs:
    by_year[r['year']][r['cat']] += 1

paper_table1 = {
    1990: (39.8, 52.3, 7.8, 3863),
    1995: (41.7, 52.4, 5.9, 4270),
    2000: (41.5, 51.2, 7.4, 6363),
    2005: (48.4, 45.7, 5.9, 5905),
    2010: (55.5, 37.1, 7.4, 7014),
    2015: (57.8, 33.0, 9.2, 8151),
    2020: (60.4, 32.4, 7.2, 9801),
    2024: (65.0, 28.4, 6.5, 5086),
}

for year in sorted(by_year):
    d = by_year[year]
    n = sum(d.values())
    art = 100 * d['article'] / n
    book = 100 * d['book'] / n
    other = 100 * d['other'] / n
    flag = "  ← paper:" + f" {paper_table1.get(year, ('?','?','?','?'))[0]}% art, {paper_table1.get(year, ('?',))[1]}% book" if year in paper_table1 else ""
    print(f"{year:>6} {art:>9.1f}% {book:>7.1f}% {other:>6.1f}% {n:>7}{flag}")


# ── FIGURE 2 / TABLE 2: % journal articles by subfield ──────────────────────
print("\n=== TABLE 2: % Journal Article References by Subfield & Year ===")

subfield_map = {
    'AmerPol': 'American Politics',
    'CompPol': 'Comparative Politics',
    'IR': 'International Relations',
    'PolTheory': 'Political Theory',
    'PubAdm': 'Public Admin',
    'Other': 'Other',
}

by_sub_year = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
for r in refs:
    if r['subfield'] and r['subfield'] != '':
        by_sub_year[r['subfield']][r['year']][r['cat']] += 1

subfields = ['AmerPol', 'CompPol', 'IR', 'PolTheory', 'PubAdm']
years = sorted(paper_table1.keys())

print(f"{'Year':>6} " + " ".join(f"{s[:8]:>10}" for s in subfields))
for year in years:
    pcts = []
    for sf in subfields:
        d = by_sub_year[sf][year]
        n = sum(d.values())
        pcts.append(100 * d['article'] / n if n else float('nan'))
    print(f"{year:>6} " + " ".join(f"{p:>9.1f}%" for p in pcts))


# ── AGE ANALYSIS (original question) ────────────────────────────────────────
print("\n=== AGE ANALYSIS: How old are cited books vs cited articles? ===")
print("(age = year of citing article minus publication year of cited work)")

age_data = defaultdict(list)
for r in refs:
    if r['cat'] in ('book', 'article') and r['pub_year'] and r['year']:
        try:
            age = int(r['year']) - int(r['pub_year'])
            if 0 <= age <= 100:  # sanity check
                age_data[r['cat']].append(age)
        except (ValueError, TypeError):
            pass

for cat in ('book', 'article'):
    ages = age_data[cat]
    ages_sorted = sorted(ages)
    n = len(ages)
    mean = sum(ages) / n
    median = ages_sorted[n // 2]
    p25 = ages_sorted[n // 4]
    p75 = ages_sorted[3 * n // 4]
    pct_recent_5 = 100 * sum(1 for a in ages if a <= 5) / n
    pct_recent_10 = 100 * sum(1 for a in ages if a <= 10) / n
    pct_recent_2 = 100 * sum(1 for a in ages if a <= 2) / n
    print(f"\n  {cat.upper()} citations (n={n:,})")
    print(f"    Mean age:   {mean:.1f} years")
    print(f"    Median age: {median} years")
    print(f"    25th–75th:  {p25}–{p75} years")
    print(f"    % cited within 2 years of publication:  {pct_recent_2:.1f}%")
    print(f"    % cited within 5 years of publication:  {pct_recent_5:.1f}%")
    print(f"    % cited within 10 years of publication: {pct_recent_10:.1f}%")

print("\n--- Age distribution (% of citations, by age bucket) ---")
buckets = [(0, 2), (3, 5), (6, 10), (11, 20), (21, 50), (51, 100)]
print(f"{'Age range':>12} {'Books%':>8} {'Articles%':>10}")
for lo, hi in buckets:
    b = sum(1 for a in age_data['book'] if lo <= a <= hi)
    a = sum(1 for a in age_data['article'] if lo <= a <= hi)
    bp = 100 * b / len(age_data['book'])
    ap = 100 * a / len(age_data['article'])
    print(f"  {lo:>2}–{hi:<3} yrs  {bp:>7.1f}% {ap:>9.1f}%")


# ── AGE ANALYSIS BY YEAR: are newer citations getting newer? ────────────────
print("\n--- Mean age of cited books vs articles, by citing year ---")
age_by_year = defaultdict(lambda: defaultdict(list))
for r in refs:
    if r['cat'] in ('book', 'article') and r['pub_year'] and r['year']:
        try:
            age = int(r['year']) - int(r['pub_year'])
            if 0 <= age <= 100:
                age_by_year[r['year']][r['cat']].append(age)
        except (ValueError, TypeError):
            pass

print(f"{'Year':>6} {'Book mean age':>14} {'Article mean age':>17} {'Book n':>8} {'Article n':>10}")
for year in sorted(age_by_year):
    b = age_by_year[year]['book']
    a = age_by_year[year]['article']
    bm = sum(b)/len(b) if b else float('nan')
    am = sum(a)/len(a) if a else float('nan')
    print(f"{year:>6} {bm:>14.1f} {am:>17.1f} {len(b):>8} {len(a):>10}")
